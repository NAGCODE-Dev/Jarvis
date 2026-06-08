from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
import uuid

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from qdrant_client import QdrantClient
from qdrant_client.http import models as qm

from jarvis.config import settings
from jarvis.ollama_client import OllamaClient
from jarvis.schemas import SearchResult


SUPPORTED_SUFFIXES = {".md", ".txt", ".pdf", ".docx", ".xlsx", ".csv"}


@dataclass
class IndexedChunk:
    point_id: str
    text: str
    metadata: dict[str, str]
    vector: list[float]


class KnowledgeService:
    def __init__(self, ollama: OllamaClient) -> None:
        self.ollama = ollama
        self.mode = "remote"
        self.client = self._build_client()

    def _build_client(self) -> QdrantClient:
        try:
            client = QdrantClient(url=settings.qdrant_url, check_compatibility=False)
            client.get_collections()
            self.mode = "remote"
            return client
        except Exception:
            settings.qdrant_local_path.mkdir(parents=True, exist_ok=True)
            self.mode = "local"
            return QdrantClient(path=str(settings.qdrant_local_path))

    def ensure_collection(self, vector_size: int | None = None, force_recreate: bool = False) -> None:
        collection = settings.qdrant_collection
        if vector_size is None:
            sample = self.ollama.embed("jarvis vector dimension probe")[0]
            vector_size = len(sample)
        existing = {c.name for c in self.client.get_collections().collections}
        if collection in existing:
            info = self.client.get_collection(collection)
            current_size = self._collection_vector_size(info)
            if current_size == vector_size:
                return
            if not force_recreate:
                raise ValueError(
                    f"Qdrant collection '{collection}' has vector size {current_size}, "
                    f"but current embeddings use size {vector_size}. "
                    "Re-run indexing with force=true to recreate the collection."
                )
            self.client.delete_collection(collection)
        self.client.create_collection(
            collection_name=collection,
            vectors_config=qm.VectorParams(size=vector_size, distance=qm.Distance.COSINE),
        )

    def health(self) -> dict[str, str | int]:
        collections = self.client.get_collections().collections
        return {
            "status": "ok",
            "collections": len(collections),
            "mode": self.mode,
        }

    def index_domains(self, domains: list[str] | None = None, force: bool = False) -> dict[str, int]:
        targets = []
        root = settings.knowledge_dir
        if domains:
            targets = [root / domain for domain in domains]
        else:
            targets = [path for path in root.iterdir() if path.is_dir()]

        indexed_files = 0
        indexed_chunks = 0
        vector_size = len(self.ollama.embed("jarvis vector dimension probe")[0])
        self.ensure_collection(vector_size=vector_size, force_recreate=force)
        for directory in targets:
            for file_path in directory.rglob("*"):
                if file_path.suffix.lower() not in SUPPORTED_SUFFIXES or not file_path.is_file():
                    continue
                chunks = self._file_to_chunks(file_path, force=force)
                if not chunks:
                    continue
                points = [
                    qm.PointStruct(id=chunk.point_id, vector=chunk.vector, payload={**chunk.metadata, "text": chunk.text})
                    for chunk in chunks
                ]
                self.client.upsert(collection_name=settings.qdrant_collection, points=points)
                indexed_files += 1
                indexed_chunks += len(chunks)
        return {"indexed_files": indexed_files, "indexed_chunks": indexed_chunks}

    def search(self, query: str, domain: str | None = None, top_k: int = 5, score_threshold: float | None = None) -> list[SearchResult]:
        vector = self.ollama.embed(query)[0]
        self.ensure_collection(vector_size=len(vector), force_recreate=False)
        query_filter = None
        if domain:
            query_filter = qm.Filter(
                must=[qm.FieldCondition(key="domain", match=qm.MatchValue(value=domain))]
            )
        results = self.client.query_points(
            collection_name=settings.qdrant_collection,
            query=vector,
            limit=min(top_k, settings.max_search_results),
            query_filter=query_filter,
            score_threshold=score_threshold,
        )
        return [
            SearchResult(
                score=point.score,
                text=point.payload.get("text", ""),
                metadata={k: v for k, v in point.payload.items() if k != "text"},
            )
            for point in results.points
        ]

    def _file_to_chunks(self, path: Path, force: bool = False) -> list[IndexedChunk]:
        text = self._extract_text(path).strip()
        if not text:
            return []
        checksum = sha256(text.encode("utf-8")).hexdigest()
        chunks = self._chunk_text(text)
        embeddings = self.ollama.embed(chunks)
        domain = path.relative_to(settings.knowledge_dir).parts[0]
        indexed_at = datetime.now(UTC).replace(microsecond=0).isoformat()
        results: list[IndexedChunk] = []
        for index, (chunk, vector) in enumerate(zip(chunks, embeddings, strict=False)):
            point_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"{path}:{checksum}:{index}"))
            metadata = {
                "source_path": str(path.relative_to(settings.data_dir)),
                "domain": domain,
                "title": path.stem,
                "checksum": checksum,
                "indexed_at": indexed_at,
                "chunk_index": str(index),
            }
            results.append(IndexedChunk(point_id=point_id, text=chunk, metadata=metadata, vector=vector))
        return results

    def _extract_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".md", ".txt"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".pdf":
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        if suffix == ".docx":
            document = Document(str(path))
            return "\n".join(paragraph.text for paragraph in document.paragraphs)
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", errors="ignore") as handle:
                return handle.read()
        if suffix == ".xlsx":
            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in workbook.worksheets:
                rows.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(value) for value in row if value is not None]
                    if cells:
                        rows.append(" | ".join(cells))
            return "\n".join(rows)
        return ""

    def _chunk_text(self, text: str, chunk_size: int = 1200, overlap: int = 200) -> list[str]:
        text = " ".join(text.split())
        if len(text) <= chunk_size:
            return [text]
        chunks: list[str] = []
        start = 0
        while start < len(text):
            end = min(len(text), start + chunk_size)
            chunks.append(text[start:end])
            if end == len(text):
                break
            start = max(0, end - overlap)
        return chunks

    def _collection_vector_size(self, info: object) -> int | None:
        config = getattr(info, "config", None)
        params = getattr(config, "params", None)
        vectors = getattr(params, "vectors", None)
        if hasattr(vectors, "size"):
            return int(vectors.size)
        return None
